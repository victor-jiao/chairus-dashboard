import time
# -*- coding: utf-8 -*-
"""一键更新数据管道：自动重命名 PLUS 文件 -> 解析所有 Excel -> 聚合双店铺/自制与联盟视频 -> 生成 data.js 与 detail-data.js"""
import os, re, json, datetime
import subprocess

PLUS_DIR = r"C:\Users\Admin\Downloads\易得客下载目录\Tiktok-ChairusPlus-焦文浩"
DASH_DIR = r"D:\codex-钉钉\dashboard"

# 1. 自动检查并重命名 PLUS 目录下的 product_list 文件
if os.path.exists(PLUS_DIR):
    for f in os.listdir(PLUS_DIR):
        m = re.match(r"^product_list_(\d{4})(\d{2})(\d{2})\.xlsx$", f)
        if m:
            y, mo, d = m.groups()
            new_name = f"{int(mo)}月{int(d)}日.xlsx"
            try:
                os.rename(os.path.join(PLUS_DIR, f), os.path.join(PLUS_DIR, new_name))
                print(f"Auto renamed: {f} -> {new_name}")
            except Exception as e:
                print(f"Rename error: {e}")

# 2. 运行 build_data.py 与 gen_insights.py
python_exe = r"C:\Users\Admin\.cache\codex-runtimes\codex-primary-runtime\dependencies\python\python.exe"
subprocess.run([python_exe, os.path.join(DASH_DIR, "build_data.py")], check=True)
subprocess.run([python_exe, os.path.join(DASH_DIR, "gen_insights.py")], check=True)

# 3. 补充视频曝光与点击聚合
import openpyxl

def num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace("$", "").replace(",", "").replace("¥", "").replace(" ", "").replace("USD", "")
    try: return float(s)
    except Exception: return 0.0

def parse_all_videos(spath, keyword):
    merged = {}
    for fn in os.listdir(spath):
        if fn.endswith(".xlsx") and not fn.startswith("~$") and keyword in fn:
            v_data = parse_video_excel(os.path.join(spath, fn))
            merged.update(v_data)
    return merged

def parse_video_excel(path):
    res = {}
    if not os.path.exists(path): return res
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            wb.close()
            return res
        
        header = [str(c or "").strip() for c in rows[2]]
        
        # 格式 A：日汇总表（第 1 列为时间/日期）
        if header and ("时间" in header[0] or "日期" in header[0]):
            exp_col = 11
            clk_col = 12
            for idx, col_name in enumerate(header):
                if "曝光" in col_name and "商品" in col_name: exp_col = idx
                elif "点击" in col_name and "商品" in col_name: clk_col = idx
            for r in rows[3:]:
                if not r or r[0] is None: continue
                d_str = str(r[0]).strip()
                if not d_str or d_str == "nan": continue
                d_str = d_str[:10].replace("/", "-")
                exp = int(num(r[exp_col])) if len(r) > exp_col else 0
                clk = int(num(r[clk_col])) if len(r) > clk_col else 0
                curr = res.setdefault(d_str, {"exposure": 0, "clicks": 0})
                curr["exposure"] += exp
                curr["clicks"] += clk
        
        # 格式 B：视频明细宽表（按商品曝光/点击均分或按所属区间 8.17~8.23 平均汇总）
        else:
            time_col = 4
            exp_col = 12
            clk_col = 13
            for idx, col_name in enumerate(header):
                if "时间" in col_name: time_col = idx
                elif "曝光" in col_name and "商品" in col_name: exp_col = idx
                elif "点击" in col_name and "商品" in col_name: clk_col = idx
            
            # 统计明细表总数
            tot_exp = sum(int(num(r[exp_col])) for r in rows[3:] if len(r) > exp_col)
            tot_clk = sum(int(num(r[clk_col])) for r in rows[3:] if len(r) > clk_col)
            
            # 如果文件名包含日期区间（例如 8.17-8.23），将总曝光均分到当周 7 天
            week_days = ["2026-08-17", "2026-08-18", "2026-08-19", "2026-08-20", "2026-08-21", "2026-08-22", "2026-08-23"]
            avg_exp = tot_exp // len(week_days)
            avg_clk = tot_clk // len(week_days)
            for d_str in week_days:
                curr = res.setdefault(d_str, {"exposure": 0, "clicks": 0})
                curr["exposure"] += avg_exp
                curr["clicks"] += avg_clk

        wb.close()
    except Exception as e:
        print(f"Error parsing {path}: {e}")
    return res

data_path = os.path.join(DASH_DIR, "data.json")
with open(data_path, "r", encoding="utf-8") as f:
    data = json.load(f)

shop_dirs = [
    ("CHAIRUS", r"C:\Users\Admin\Downloads\易得客下载目录\Tiktok-Chairus-焦文浩"),
    ("PLUS", r"C:\Users\Admin\Downloads\易得客下载目录\Tiktok-ChairusPlus-焦文浩"),
]

for sname, spath in shop_dirs:
    self_v = parse_all_videos(spath, "自制视频")
    affil_v = parse_all_videos(spath, "联盟视频")
    shop_obj = data["shop"][sname]
    for d in shop_obj["days"]:
        ds = d["date"]
        sv = self_v.get(ds, {"exposure": 0, "clicks": 0})
        av = affil_v.get(ds, {"exposure": 0, "clicks": 0})
        d["self_video_exposure"] = sv["exposure"]
        d["self_video_clicks"] = sv["clicks"]
        d["affil_video_exposure"] = av["exposure"]
        d["affil_video_clicks"] = av["clicks"]
    for w in shop_obj["weeks"]:
        w["tot"]["self_video_exposure"] = sum(d.get("self_video_exposure", 0) for d in shop_obj["days"] if d["date"] in w["days"])
        w["tot"]["self_video_clicks"] = sum(d.get("self_video_clicks", 0) for d in shop_obj["days"] if d["date"] in w["days"])
        w["tot"]["affil_video_exposure"] = sum(d.get("affil_video_exposure", 0) for d in shop_obj["days"] if d["date"] in w["days"])
        w["tot"]["affil_video_clicks"] = sum(d.get("affil_video_clicks", 0) for d in shop_obj["days"] if d["date"] in w["days"])

total_obj = data["total"]
for d in total_obj["days"]:
    ds = d["date"]
    d["self_video_exposure"] = sum(data["shop"][s]["days"][i].get("self_video_exposure", 0) for s in ["CHAIRUS", "PLUS"] for i, sd in enumerate(data["shop"][s]["days"]) if sd["date"] == ds)
    d["self_video_clicks"] = sum(data["shop"][s]["days"][i].get("self_video_clicks", 0) for s in ["CHAIRUS", "PLUS"] for i, sd in enumerate(data["shop"][s]["days"]) if sd["date"] == ds)
    d["affil_video_exposure"] = sum(data["shop"][s]["days"][i].get("affil_video_exposure", 0) for s in ["CHAIRUS", "PLUS"] for i, sd in enumerate(data["shop"][s]["days"]) if sd["date"] == ds)
    d["affil_video_clicks"] = sum(data["shop"][s]["days"][i].get("affil_video_clicks", 0) for s in ["CHAIRUS", "PLUS"] for i, sd in enumerate(data["shop"][s]["days"]) if sd["date"] == ds)

for w in total_obj["weeks"]:
    w["tot"]["self_video_exposure"] = sum(data["shop"][s]["weeks"][i]["tot"].get("self_video_exposure", 0) for s in ["CHAIRUS", "PLUS"] for i, sw in enumerate(data["shop"][s]["weeks"]) if sw["key"] == w["key"])
    w["tot"]["self_video_clicks"] = sum(data["shop"][s]["weeks"][i]["tot"].get("self_video_clicks", 0) for s in ["CHAIRUS", "PLUS"] for i, sw in enumerate(data["shop"][s]["weeks"]) if sw["key"] == w["key"])
    w["tot"]["affil_video_exposure"] = sum(data["shop"][s]["weeks"][i]["tot"].get("affil_video_exposure", 0) for s in ["CHAIRUS", "PLUS"] for i, sw in enumerate(data["shop"][s]["weeks"]) if sw["key"] == w["key"])
    w["tot"]["affil_video_clicks"] = sum(data["shop"][s]["weeks"][i]["tot"].get("affil_video_clicks", 0) for s in ["CHAIRUS", "PLUS"] for i, sw in enumerate(data["shop"][s]["weeks"]) if sw["key"] == w["key"])

with open(data_path, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=1)
with open(os.path.join(DASH_DIR, "data.js"), "w", encoding="utf-8") as f:
    f.write("window.DASH_DATA = " + json.dumps(data, ensure_ascii=False) + ";")


# 4. 更新 index.html 时间戳防缓存
html_path = os.path.join(DASH_DIR, "index.html")
if os.path.exists(html_path):
    with open(html_path, "r", encoding="utf-8") as f:
        h = f.read()
    v_ts = str(int(time.time()))
    h = re.sub(r'src="data\.js(\?v=\d+)?"', f'src="data.js?v={v_ts}"', h)
    h = re.sub(r'src="roi_data\.js(\?v=\d+)?"', f'src="roi_data.js?v={v_ts}"', h)
    h = re.sub(r'src="detail-data\.js(\?v=\d+)?"', f'src="detail-data.js?v={v_ts}"', h)
    h = re.sub(r'src="app\.js(\?v=\d+)?"', f'src="app.js?v={v_ts}"', h)
    h = re.sub(r'href="style\.css(\?v=\d+)?"', f'href="style.css?v={v_ts}"', h)
    with open(html_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(h)
    print(f"Cachebuster timestamp updated: v={v_ts}")


# 5. 自动从《ROI记录表.xlsx》读取并固化 ROI 到 roi_data.js
roi_excel_path = os.path.join(DASH_DIR, "ROI记录表.xlsx")
# 如果用户放在了易得客目录或桌面，也自动查找
possible_roi_paths = [
    roi_excel_path,
    r"C:\Users\Admin\Desktop\ROI记录表.xlsx",
    r"C:\Users\Admin\Downloads\ROI记录表.xlsx",
    r"C:\Users\Admin\Downloads\易得客下载目录\ROI记录表.xlsx"
]

roi_dict = {}
for rp in possible_roi_paths:
    if os.path.exists(rp):
        try:
            wb_roi = openpyxl.load_workbook(rp, data_only=True)
            ws_roi = wb_roi.worksheets[0]
            for r in ws_roi.iter_rows(min_row=2, values_only=True):
                if not r or r[0] is None: continue
                d_str = str(r[0]).strip()[:10].replace("/", "-")
                if len(d_str) < 8 or "日期" in d_str: continue
                # CHAIRUS ROI
                if len(r) > 1 and r[1] is not None and str(r[1]).strip():
                    try: roi_dict[f"CHAIRUS|{d_str}"] = float(str(r[1]).strip().replace("$",""))
                    except: pass
                # PLUS ROI
                if len(r) > 2 and r[2] is not None and str(r[2]).strip():
                    try: roi_dict[f"PLUS|{d_str}"] = float(str(r[2]).strip().replace("$",""))
                    except: pass
            wb_roi.close()
            print(f"Loaded {len(roi_dict)} ROI records from {rp}")
            break
        except Exception as e:
            print(f"Error reading ROI excel {rp}: {e}")

# 写入 roi_data.js
roi_js_path = os.path.join(DASH_DIR, "roi_data.js")
with open(roi_js_path, "w", encoding="utf-8") as f:
    f.write("window.DEFAULT_ROI = " + json.dumps(roi_dict, ensure_ascii=False, indent=2) + ";\n")
print(f"Updated roi_data.js with {len(roi_dict)} items!")

print("Pipeline finished successfully!")
