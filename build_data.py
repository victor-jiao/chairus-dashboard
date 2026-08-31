# -*- coding: utf-8 -*-
"""TikTok 数据看板 - 双店铺数据管道：Excel -> data.json / data.js / detail-data.js"""
import os, re, json, datetime
import openpyxl

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
SHOP_DIRS = [
    ("CHAIRUS", r"C:\Users\Admin\Downloads\易得客下载目录\Tiktok-Chairus-焦文浩"),
    ("PLUS", r"C:\Users\Admin\Downloads\易得客下载目录\Tiktok-ChairusPlus-焦文浩"),
    ("HOME", r"C:\Users\Admin\Downloads\易得客下载目录\ChairusHome子账号-焦文浩"),
]
GMV_PATH = r"C:\Users\Admin\Downloads\月度GMV完成度.xlsx"
DEFAULT_GMV_PATH = r"D:\codex-钉钉\月度GMV完成度_提取.json"

def num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s == "" or s.lower() in ("-", "n/a"): return 0.0
    s = s.replace("$", "").replace(",", "").replace("¥", "").replace(" ", "").replace("USD", "")
    try: return float(s)
    except Exception: return 0.0

def date_from_daily_name(fn):
    m = re.search(r"(\d+)月(\d+)日", fn)
    if m: return datetime.date(2026, int(m.group(1)), int(m.group(2)))
    m2 = re.match(r"product_list_(\d{4})(\d{2})(\d{2})", fn)
    if m2: return datetime.date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
    return None

def sku_extract(sku):
    s = str(sku).strip()
    for pre in ("TKYJ-", "TKJN-"):
        if s.startswith(pre):
            s = s[len(pre):]; break
    if not s: return ""
    if s[0] == "A":
        m = re.match(r"^A[A-Za-z]{3,6}", s)
        if m:
            n = len(m.group(0))
            if n == 4: return s[3:10] if len(s) >= 10 else s[3:]
            if n >= 5: return s[4:10] if len(s) >= 10 else s[4:]
    return s

def month_first_monday(M):
    d1 = datetime.date(2026, M, 1)
    mon = d1 - datetime.timedelta(days=d1.weekday())
    inM = sum(1 for i in range(7) if (mon + datetime.timedelta(days=i)).month == M)
    if inM < 4: mon = mon + datetime.timedelta(days=7)
    return mon

def week_of_date(d):
    mon = d - datetime.timedelta(days=d.weekday())
    sunday = mon + datetime.timedelta(days=6)
    cnt = {}
    for i in range(7):
        m = (mon + datetime.timedelta(days=i)).month
        cnt[m] = cnt.get(m, 0) + 1
    owner = max(cnt, key=lambda m: cnt[m])
    fmon = month_first_monday(owner)
    wn = (mon - fmon).days // 7 + 1
    return owner, wn, mon, sunday

def load_product_map(path):
    pm = {}
    if not os.path.exists(path): return pm
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb["Template"] if "Template" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx, hdr = None, None
    for i, r in enumerate(rows):
        vals = [str(x).strip() if x is not None else "" for x in r]
        if "商品 ID" in vals:
            hdr_idx, hdr = i, vals
            break
    if hdr is None: return pm
    def col(name):
        for j, h in enumerate(hdr):
            if h == name: return j
        return None
    i_pid, i_cat, i_sku = col("商品 ID"), col("类目"), col("商家 SKU")
    if i_pid is None: return pm
    for r in rows[hdr_idx + 3:]:
        if r[i_pid] is None: continue
        pid_s = str(r[i_pid]).strip()
        if not pid_s or not pid_s.isdigit(): continue
        cat_s = str(r[i_cat]).strip() if i_cat is not None and r[i_cat] is not None else ""
        clean = re.sub(r"\s*\(\d+\)\s*$", "", cat_s).strip()
        sku_s = str(r[i_sku]).strip() if i_sku is not None and r[i_sku] is not None else ""
        pm[pid_s] = {"category": clean, "category_raw": cat_s, "sku": sku_s}
    wb.close()
    return pm

DAILY_COLS = {
    "gmv": 4, "live_gmv": 5, "video_gmv": 8, "affil_gmv": 11, "card_gmv": 18,
    "orders": 19, "qty": 21, "customers": 22, "exposure": 24, "clicks": 25,
    "add_cart": 27, "mall_exposure": 42, "mall_clicks": 43, "mall_gmv": 49,
    "refund": 40, "refund_qty": 41,
}

def build_detail_schema(ws):
    r3 = [c.value for c in ws[3]]
    r4 = [c.value for c in ws[4]]
    keep = []
    for i in range(len(r4)):
        if i in (0, 1, 2, 3): continue
        h3 = str(r3[i]).strip() if r3[i] is not None else ""
        h4 = str(r4[i]).strip() if r4[i] is not None else ""
        merged = (h3 + h4) if h3 else h4
        if "直播" in h3 or "直播" in h4: continue
        if "含税 GMV" in merged or "税费" in merged: continue
        if "去重" in merged: continue
        keep.append((i, merged))
    return keep

def fmt_detail_val(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return v
    s = str(v).strip()
    if s == "": return 0
    if s.endswith("%"):
        try: return round(float(s[:-1].replace(",", "")), 4)
        except Exception: return 0
    if re.match(r"^[\d,.]+$", s):
        try: return float(s.replace(",", ""))
        except Exception: return s
    if "$" in s:
        try: return round(float(s.replace("$", "").replace(",", "")), 2)
        except Exception: return s
    return s

def parse_order_time(v):
    if v is None: return None
    s = str(v).strip()
    for f in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try: return datetime.datetime.strptime(s, f).date()
        except Exception: pass
    return None

def load_orders(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    def idx(name):
        for i, h in enumerate(hdr):
            if h.lower() == name.lower(): return i
        return None
    i_status, i_cancel = idx("Order Status"), idx("Cancelation/Return Type")
    i_sku, i_qty, i_cat, i_channel, i_time = idx("Seller SKU"), idx("Quantity"), idx("Product Category"), idx("Order Channel"), idx("Created Time")
    i_gmv = idx("SKU Subtotal After Discount")
    out = []
    for r in rows[2:]:
        if i_sku is None or r[i_sku] is None: continue
        status = str(r[i_status]).strip() if i_status is not None and r[i_status] else ""
        cancel = str(r[i_cancel]).strip() if i_cancel is not None and r[i_cancel] else ""
        if cancel: continue
        if any(k in status for k in ("取消", "退货", "退款", "Cancel", "Return", "Refund")): continue
        out.append({
            "date": parse_order_time(r[i_time]).isoformat() if (i_time is not None and parse_order_time(r[i_time])) else None,
            "sku": str(r[i_sku]).strip(), "qty": int(num(r[i_qty])),
            "gmv": round(num(r[i_gmv]), 2) if i_gmv is not None else 0.0,
            "code": sku_extract(str(r[i_sku]).strip()),
            "category": str(r[i_cat]).strip() if i_cat is not None and r[i_cat] else "",
            "channel": str(r[i_channel]).strip() if i_channel is not None and r[i_channel] else "",
        })
    return out

def process_shop(name, path):
    pm = load_product_map(os.path.join(path, "商品信息.xlsx"))
    daily_files = []
    for fn in os.listdir(path):
        if not fn.endswith(".xlsx") or fn.startswith("~$") or "视频" in fn: continue
        if "订单" in fn: continue
        d = date_from_daily_name(fn)
        if d: daily_files.append((d, os.path.join(path, fn)))
    daily_files.sort(key=lambda x: x[0])

    days, day_cat = [], {}
    detail_days = []
    for d, fp in daily_files:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=False)
        ws = wb.worksheets[0]
        raw = [r for r in ws.iter_rows(min_row=5, values_only=True) if (r[1] is not None or r[0] is not None)]
        schema = build_detail_schema(ws)
        headers = ["品类"] + [h for _, h in schema]
        rows = []
        tot = {k: 0.0 for k in DAILY_COLS}
        cat_agg = {}
        for r in raw:
            if len(r) < 30: continue
            pid = str(r[1]).strip() if r[1] is not None else ""
            if not pid: continue
            info = pm.get(pid)
            code = sku_extract(info["sku"]) if info and info["sku"] else ""
            cat = info["category"] if info else "未分类"
            g = {k: num(r[i]) if i < len(r) else 0.0 for k, i in DAILY_COLS.items()}
            for k in tot: tot[k] += g[k]
            ca = cat_agg.setdefault(cat, {k: 0.0 for k in ("gmv","orders","qty","exposure","clicks","add_cart","refund")})
            for k in ca: ca[k] += g[k]
            if not code: continue
            row = [code] + [fmt_detail_val(r[i] if i < len(r) else None) for i, _ in schema]
            rows.append(row)
        ds = d.strftime("%Y-%m-%d")
        days.append({
            "date": ds, "label": f"{d.month}月{d.day}日",
            "gmv": round(tot["gmv"], 2),
            "sources": {"mall": round(tot["mall_gmv"],2), "video": round(tot["video_gmv"],2),
                        "affil": round(tot["affil_gmv"],2), "card": round(tot["card_gmv"],2),
                        "live": round(tot["live_gmv"],2)},
            "orders": int(tot["orders"]), "qty": int(tot["qty"]), "customers": int(tot["customers"]),
            "exposure": int(tot["exposure"]), "clicks": int(tot["clicks"]), "add_cart": int(tot["add_cart"]),
            "refund": round(tot["refund"],2), "refund_qty": int(tot["refund_qty"]),
            "categories": sorted(cat_agg.keys()),
        })
        for cat, m in cat_agg.items():
            day_cat[(ds, cat)] = m
        detail_days.append({"date": ds, "label": f"{d.month}月{d.day}日", "headers": headers, "rows": rows})
        wb.close()

    # 读取自制视频数据
    self_video_map = {}
    self_video_file = os.path.join(path, "6.22-8.16自制视频.xlsx")
    if os.path.exists(self_video_file):
        try:
            vwb = openpyxl.load_workbook(self_video_file, data_only=True, read_only=False)
            vws = vwb.worksheets[0]
            for vrow in vws.iter_rows(min_row=4, values_only=True):
                if not vrow or vrow[0] is None: continue
                vdate_str = str(vrow[0]).strip()
                if not vdate_str or vdate_str == "nan": continue
                vexp = int(num(vrow[11])) if len(vrow) > 11 else 0
                vclk = int(num(vrow[12])) if len(vrow) > 12 else 0
                self_video_map[vdate_str] = {"exposure": vexp, "clicks": vclk}
            vwb.close()
        except Exception as ve:
            print("Self video load error:", ve)

    for x in days:
        sv = self_video_map.get(x["date"], {"exposure": 0, "clicks": 0})
        x["video_exposure"] = sv["exposure"]
        x["video_clicks"] = sv["clicks"]

    week_map = {}
    for x in days:
        d = datetime.date.fromisoformat(x["date"])
        om, wn, mon, sun = week_of_date(d)
        key = f"{om}-{wn}"
        wk = week_map.setdefault(key, {"key": key, "month": om, "week": wn,
            "monday": mon.strftime("%Y-%m-%d"), "sunday": sun.strftime("%Y-%m-%d"),
            "days": [], "tot": {k: 0.0 for k in ("gmv","orders","qty","exposure","clicks","add_cart","refund","video_exposure","video_clicks")},
            "cat": {}, "video_gmv": 0.0})
        wk["days"].append(x["date"])
        for k in wk["tot"]:
            wk["tot"][k] += x.get(k, 0.0) if k in ("gmv","refund") else float(x.get(k, 0))
        wk["video_gmv"] = round(wk.get("video_gmv", 0.0) + x["sources"]["video"], 2)
        for cat in x["categories"]:
            cm = day_cat.get((x["date"], cat))
            if cm is None: continue
            cc = wk["cat"].setdefault(cat, {"gmv":0.0,"orders":0,"qty":0,"exposure":0,"clicks":0,"add_cart":0})
            for k in cc: cc[k] += cm[k]
    weeks = sorted([w for w in week_map.values() if len(w["days"]) == 7], key=lambda w: (w["month"], w["week"]))

    order_files = [os.path.join(path, fn) for fn in os.listdir(path) if fn.endswith(".xlsx") and "订单" in fn]
    order_rows = []
    for p in order_files:
        order_rows.extend(load_orders(p))

    sku_qty, sku_gmv, cat_qty, cat_gmv, channel_qty = {}, {}, {}, {}, {}
    sku_cat_map = {}
    cat_sku_map = {}

    for o in order_rows:
        if not o["code"]: continue
        code = o["code"]
        cat = o["category"] or "未分类"
        sku_qty[code] = sku_qty.get(code, 0) + o["qty"]
        sku_gmv[code] = sku_gmv.get(code, 0) + o["gmv"]
        if cat:
            cat_qty[cat] = cat_qty.get(cat, 0) + o["qty"]
            cat_gmv[cat] = cat_gmv.get(cat, 0.0) + o["gmv"]
            sku_cat_map[code] = cat
            cat_sku_map.setdefault(cat, set()).add(code)
        if o["channel"]: channel_qty[o["channel"]] = channel_qty.get(o["channel"], 0) + o["qty"]

    # 周维度：各 SKU & 各品类的周销量 / 周 GMV / 周订单数
    week_sku = {}
    week_sku_gmv = {}
    week_sku_orders = {}
    week_cat_qty = {}
    week_cat_gmv = {}
    week_cat_orders = {}
    week_info = {}

    for o in order_rows:
        if not o["date"] or not o["code"]: continue
        om, wn, mon, sun = week_of_date(datetime.date.fromisoformat(o["date"]))
        key = f"{om}-{wn}"
        code = o["code"]
        cat = o["category"] or "未分类"

        week_info.setdefault(key, {"monday": mon.strftime("%Y-%m-%d"), "sunday": sun.strftime("%Y-%m-%d")})

        # SKU 统计
        ws = week_sku.setdefault(key, {})
        ws[code] = ws.get(code, 0) + o["qty"]

        wg = week_sku_gmv.setdefault(key, {})
        wg[code] = round(wg.get(code, 0.0) + o["gmv"], 2)

        wo = week_sku_orders.setdefault(key, {})
        wo[code] = wo.get(code, 0) + 1

        # 品类统计
        wcq = week_cat_qty.setdefault(key, {})
        wcq[cat] = wcq.get(cat, 0) + o["qty"]

        wcg = week_cat_gmv.setdefault(key, {})
        wcg[cat] = round(wcg.get(cat, 0.0) + o["gmv"], 2)

        wco = week_cat_orders.setdefault(key, {})
        wco[cat] = wco.get(cat, 0) + 1

    sorted_weeks = sorted(week_info.keys())
    top_skus = [k for k, _ in sorted(sku_qty.items(), key=lambda x: x[1], reverse=True)[:10]]
    top_gmv_skus = [k for k, _ in sorted(sku_gmv.items(), key=lambda x: x[1], reverse=True)[:10]]
    sorted_cats = [k for k, _ in sorted(cat_qty.items(), key=lambda x: x[1], reverse=True)]

    # 构建结构化的品类下 SKU 列表（按销量排）
    cat_skus_dict = {}
    for c in sorted_cats:
        c_skus = list(cat_sku_map.get(c, []))
        c_skus.sort(key=lambda x: sku_qty.get(x, 0), reverse=True)
        cat_skus_dict[c] = c_skus

    all_skus = list(sku_qty.keys())
    all_sku_series = {code: [week_sku.get(w, {}).get(code, 0) for w in sorted_weeks] for code in all_skus}
    all_sku_gmv_series = {code: [week_sku_gmv.get(w, {}).get(code, 0.0) for w in sorted_weeks] for code in all_skus}
    all_sku_orders_series = {code: [week_sku_orders.get(w, {}).get(code, 0) for w in sorted_weeks] for code in all_skus}

    cat_series = {c: [week_cat_qty.get(w, {}).get(c, 0) for w in sorted_weeks] for c in sorted_cats}
    cat_gmv_series = {c: [week_cat_gmv.get(w, {}).get(c, 0.0) for w in sorted_weeks] for c in sorted_cats}
    cat_orders_series = {c: [week_cat_orders.get(w, {}).get(c, 0) for w in sorted_weeks] for c in sorted_cats}

    week_sku_data = {
        "weeks": sorted_weeks,
        "week_dates": week_info,
        "skus": top_skus,
        "series": {code: all_sku_series[code] for code in top_skus},
        "all_skus": all_skus,
        "all_sku_series": all_sku_series,
        "all_sku_gmv_series": all_sku_gmv_series,
        "all_sku_orders_series": all_sku_orders_series,
        "categories": sorted_cats,
        "cat_skus": cat_skus_dict,
        "cat_series": cat_series,
        "cat_gmv_series": cat_gmv_series,
        "cat_orders_series": cat_orders_series,
        "sku_cat_map": sku_cat_map,
        "gmv_weeks": sorted_weeks,
        "gmv_dates": {k: [v["monday"], v["sunday"]] for k, v in week_info.items()},
        "gmv_skus": top_gmv_skus,
        "gmv_series": {code: all_sku_gmv_series[code] for code in top_gmv_skus},
    }

    sales_dates = {o["date"] for o in order_rows if o["date"]}
    return {
        "name": name,
        "range": {"from": days[0]["date"] if days else "", "to": days[-1]["date"] if days else ""},
        "sales_range": {"from": min(sales_dates) if sales_dates else "", "to": max(sales_dates) if sales_dates else ""},
        "days": days, "weeks": weeks,
        "sales_rank_sku": sorted([{"name": k, "qty": v} for k, v in sku_qty.items()], key=lambda x: x["qty"], reverse=True),
        "sales_rank_category": sorted([{"name": k, "qty": v} for k, v in cat_qty.items()], key=lambda x: x["qty"], reverse=True),
        "sales_rank_channel": sorted([{"name": k, "qty": v} for k, v in channel_qty.items()], key=lambda x: x["qty"], reverse=True),
        "order_rows": order_rows,
        "week_sku": week_sku_data,
        "funnel": {"exposure": sum(int(x["exposure"]) for x in days), "clicks": sum(int(x["clicks"]) for x in days),
                   "add_cart": sum(int(x["add_cart"]) for x in days), "orders": sum(int(x["orders"]) for x in days),
                   "qty": sum(int(x["qty"]) for x in days)},
        "detail_days": detail_days,
        "insights": [],
    }

def merge_total(shops):
    dmap = {}
    for s in shops:
        for d in s["days"]:
            t = dmap.setdefault(d["date"], {"date": d["date"], "label": d["label"], "sources": {k: 0.0 for k in ("mall","video","affil","card","live")},
                "orders": 0, "qty": 0, "customers": 0, "exposure": 0, "clicks": 0, "add_cart": 0,
                "refund": 0.0, "refund_qty": 0, "video_exposure": 0, "video_clicks": 0, "categories": set()})
            t["gmv"] = round(t.get("gmv", 0.0) + d["gmv"], 2)
            for k in t["sources"]: t["sources"][k] = round(t["sources"][k] + d["sources"][k], 2)
            t["orders"] += d["orders"]; t["qty"] += d["qty"]; t["customers"] += d["customers"]
            t["exposure"] += d["exposure"]; t["clicks"] += d["clicks"]; t["add_cart"] += d["add_cart"]; t["video_exposure"] += d.get("video_exposure", 0); t["video_clicks"] += d.get("video_clicks", 0)
            t["refund"] = round(t["refund"] + d["refund"], 2); t["refund_qty"] += d["refund_qty"]
            t["categories"].update(d["categories"])
    days = []
    for k in sorted(dmap):
        x = dmap[k]; x["categories"] = sorted(x["categories"]); days.append(x)

    key_sets = [set(w["key"] for w in s["weeks"]) for s in shops]
    common = set.intersection(*key_sets) if key_sets else set()
    wmap = {}
    for s in shops:
        for w in s["weeks"]:
            if w["key"] not in common: continue
            t = wmap.setdefault(w["key"], {"key": w["key"], "month": w["month"], "week": w["week"],
                "monday": w["monday"], "sunday": w["sunday"], "days": w["days"][:],
                "tot": {k: 0.0 for k in ("gmv","orders","qty","exposure","clicks","add_cart","refund","video_exposure","video_clicks")}, "cat": {}, "video_gmv": 0.0})
            for k in t["tot"]: t["tot"][k] += w["tot"][k]
            t["video_gmv"] = round(t["video_gmv"] + w["video_gmv"], 2)
            for c, m in w["cat"].items():
                cc = t["cat"].setdefault(c, {"gmv":0.0,"orders":0,"qty":0,"exposure":0,"clicks":0,"add_cart":0})
                for k in cc: cc[k] += m[k]
    weeks = [wmap[k] for k in sorted(wmap)]

    def merge_rank(shops, key):
        agg = {}
        for s in shops:
            for r in s[key]: agg[r["name"]] = agg.get(r["name"], 0) + r["qty"]
        return sorted([{"name": k, "qty": v} for k, v in agg.items()], key=lambda x: x["qty"], reverse=True)

    all_weeks_set = set()
    week_info = {}
    for s in shops:
        ws = s["week_sku"]
        all_weeks_set.update(ws["weeks"])
        for k, v in ws.get("week_dates", {}).items():
            week_info.setdefault(k, v)
    sorted_weeks = sorted(all_weeks_set)

    sku_tot_qty = {}
    sku_tot_gmv = {}
    cat_tot_qty = {}
    cat_tot_gmv = {}
    sku_cat_map = {}
    cat_sku_map = {}

    all_sku_series = {}
    all_sku_gmv_series = {}
    all_sku_orders_series = {}

    all_skus_set = set()
    for s in shops:
        all_skus_set.update(s["week_sku"].get("all_skus", []))
        for k, v in s["week_sku"].get("sku_cat_map", {}).items():
            sku_cat_map.setdefault(k, v)

    for code in all_skus_set:
        q_list = [0] * len(sorted_weeks)
        g_list = [0.0] * len(sorted_weeks)
        o_list = [0] * len(sorted_weeks)
        for s in shops:
            ws = s["week_sku"]
            w_idx_map = {w: i for i, w in enumerate(ws["weeks"])}
            sq = ws.get("all_sku_series", {}).get(code)
            sg = ws.get("all_sku_gmv_series", {}).get(code)
            so = ws.get("all_sku_orders_series", {}).get(code)
            for i, w in enumerate(sorted_weeks):
                if w in w_idx_map:
                    idx = w_idx_map[w]
                    if sq and idx < len(sq): q_list[i] += sq[idx]
                    if sg and idx < len(sg): g_list[i] = round(g_list[i] + sg[idx], 2)
                    if so and idx < len(so): o_list[i] += so[idx]
        all_sku_series[code] = q_list
        all_sku_gmv_series[code] = g_list
        all_sku_orders_series[code] = o_list
        sku_tot_qty[code] = sum(q_list)
        sku_tot_gmv[code] = round(sum(g_list), 2)
        cat = sku_cat_map.get(code, "未分类")
        cat_sku_map.setdefault(cat, set()).add(code)

    all_cats_set = set()
    for s in shops:
        all_cats_set.update(s["week_sku"].get("categories", []))

    cat_series = {}
    cat_gmv_series = {}
    cat_orders_series = {}
    for c in all_cats_set:
        q_list = [0] * len(sorted_weeks)
        g_list = [0.0] * len(sorted_weeks)
        o_list = [0] * len(sorted_weeks)
        for s in shops:
            ws = s["week_sku"]
            w_idx_map = {w: i for i, w in enumerate(ws["weeks"])}
            cq = ws.get("cat_series", {}).get(c)
            cg = ws.get("cat_gmv_series", {}).get(c)
            co = ws.get("cat_orders_series", {}).get(c)
            for i, w in enumerate(sorted_weeks):
                if w in w_idx_map:
                    idx = w_idx_map[w]
                    if cq and idx < len(cq): q_list[i] += cq[idx]
                    if cg and idx < len(cg): g_list[i] = round(g_list[i] + cg[idx], 2)
                    if co and idx < len(co): o_list[i] += co[idx]
        cat_series[c] = q_list
        cat_gmv_series[c] = g_list
        cat_orders_series[c] = o_list
        cat_tot_qty[c] = sum(q_list)
        cat_tot_gmv[c] = round(sum(g_list), 2)

    sorted_cats = sorted(all_cats_set, key=lambda c: cat_tot_qty.get(c, 0), reverse=True)
    top_skus = [k for k, _ in sorted(sku_tot_qty.items(), key=lambda x: x[1], reverse=True)[:10]]
    top_gmv_skus = [k for k, _ in sorted(sku_tot_gmv.items(), key=lambda x: x[1], reverse=True)[:10]]

    cat_skus_dict = {}
    for c in sorted_cats:
        c_skus = list(cat_sku_map.get(c, []))
        c_skus.sort(key=lambda x: sku_tot_qty.get(x, 0), reverse=True)
        cat_skus_dict[c] = c_skus

    week_sku_data = {
        "weeks": sorted_weeks,
        "week_dates": week_info,
        "skus": top_skus,
        "series": {code: all_sku_series[code] for code in top_skus},
        "all_skus": list(all_skus_set),
        "all_sku_series": all_sku_series,
        "all_sku_gmv_series": all_sku_gmv_series,
        "all_sku_orders_series": all_sku_orders_series,
        "categories": sorted_cats,
        "cat_skus": cat_skus_dict,
        "cat_series": cat_series,
        "cat_gmv_series": cat_gmv_series,
        "cat_orders_series": cat_orders_series,
        "sku_cat_map": sku_cat_map,
        "gmv_weeks": sorted_weeks,
        "gmv_dates": {k: [v.get("monday",""), v.get("sunday","")] for k, v in week_info.items()},
        "gmv_skus": top_gmv_skus,
        "gmv_series": {code: all_sku_gmv_series[code] for code in top_gmv_skus},
    }

    funnel = {"exposure": 0, "clicks": 0, "add_cart": 0, "orders": 0, "qty": 0}
    for s in shops:
        for k in funnel: funnel[k] += s["funnel"][k]
    sales_dates = set()
    for s in shops: sales_dates.update({o["date"] for o in s["order_rows"] if o["date"]})

    return {
        "name": "TOTAL", "range": {"from": days[0]["date"] if days else "", "to": days[-1]["date"] if days else ""},
        "sales_range": {"from": min(sales_dates) if sales_dates else "", "to": max(sales_dates) if sales_dates else ""},
        "days": days, "weeks": weeks,
        "sales_rank_sku": merge_rank(shops, "sales_rank_sku"),
        "sales_rank_category": merge_rank(shops, "sales_rank_category"),
        "sales_rank_channel": merge_rank(shops, "sales_rank_channel"),
        "order_rows": [o for s in shops for o in s["order_rows"]],
        "week_sku": week_sku_data,
        "funnel": funnel,
        "insights": [],
    }

def load_monthly_targets(path):
    monthly = []
    possible_paths = [
        path,
        r"C:\Users\Admin\Desktop\月度GMV完成度.xlsx",
        r"D:\codex-钉钉\月度GMV完成度.xlsx",
        r"C:\Users\Admin\Downloads\易得客下载目录\月度GMV完成度.xlsx"
    ]
    target_file = None
    for p in possible_paths:
        if os.path.exists(p):
            target_file = p
            break
            
    if target_file:
        try:
            wb = openpyxl.load_workbook(target_file, data_only=True, read_only=False)
            ws = wb.worksheets[0]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[0] is None: continue
                actual, target = num(r[1]), num(r[2])
                monthly.append({"month": str(r[0]).strip(), "actual": actual, "target": target,
                                "completion": round(actual / target * 100, 1) if target else 0})
            wb.close()
        except Exception as e:
            print("Error loading GMV targets excel:", e)
            
    if not monthly:
        # Fallback preset targets
        targets = [
            ("1月", 0, 50000), ("2月", 0, 50000), ("3月", 0, 60000),
            ("4月", 0, 80000), ("5月", 0, 100000), ("6月", 0, 150000),
            ("7月", 0, 180000), ("8月", 0, 200000), ("9月", 0, 220000),
            ("10月", 0, 250000), ("11月", 0, 300000), ("12月", 0, 350000)
        ]
        monthly = [{"month": m, "actual": a, "target": t, "completion": 0} for m, a, t in targets]
    return monthly

def compute_monthly_and_quarters(total, base_monthly):
    # 计算所有店铺各月的每日 GMV 实际求和
    month_sums = {}
    for d in total.get("days", []):
        m_num = int(d["date"].split("-")[1])
        m_key = f"{m_num}月"
        month_sums[m_key] = round(month_sums.get(m_key, 0.0) + d["gmv"], 2)
    
    monthly = []
    for m in base_monthly:
        m_key = m["month"]
        # 如果两店每日数据里有该月的实际 GMV，以所有店铺实际每日 GMV 之和为准
        actual = month_sums[m_key] if m_key in month_sums else m["actual"]
        target = m["target"]
        comp = round(actual / target * 100, 1) if target else 0
        monthly.append({"month": m_key, "actual": round(actual, 2), "target": round(target, 2), "completion": comp})

    quarters = []
    q_map = [("Q1", ["1月", "2月", "3月"]), ("Q2", ["4月", "5月", "6月"]), ("Q3", ["7月", "8月", "9月"]), ("Q4", ["10月", "11月", "12月"])]
    for qname, q_months in q_map:
        ms = [m for m in monthly if m["month"] in q_months]
        actual = sum(m["actual"] for m in ms)
        target = sum(m["target"] for m in ms)
        quarters.append({"quarter": qname, "actual": round(actual, 2), "target": round(target, 2),
                         "completion": round(actual / target * 100, 1) if target else 0, "months": [m["month"] for m in ms]})
    return monthly, quarters

def main():
    shops = [process_shop(n, p) for n, p in SHOP_DIRS]
    total = merge_total(shops)
    base_monthly = load_monthly_targets(GMV_PATH)
    monthly, quarters = compute_monthly_and_quarters(total, base_monthly)
    for s in shops:
        print(s["name"], "days", len(s["days"]), "weeks", [w["key"] for w in s["weeks"]], "orders", len(s["order_rows"]), "detail", len(s["detail_days"]))
    print("TOTAL days", len(total["days"]), "weeks", [w["key"] for w in total["weeks"]])

    data = {
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "shops": [s["name"] for s in shops],
        "total": total,
        "shop": {s["name"]: s for s in shops},
        "monthly_gmv": monthly,
        "quarters": quarters,
        "insights": [],
    }

    def write_js(path, var, obj):
        with open(path, "w", encoding="utf-8") as f:
            f.write("window." + var + " = " + json.dumps(obj, ensure_ascii=False) + ";")

    with open(os.path.join(OUT_DIR, "data.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    write_js(os.path.join(OUT_DIR, "data.js"), "DASH_DATA", data)
    write_js(os.path.join(OUT_DIR, "detail-data.js"), "DETAIL_DATA", {"shops": {s["name"]: {"days": s["detail_days"]} for s in shops}})
    print("OUTPUT ok")

if __name__ == "__main__":
    main()
